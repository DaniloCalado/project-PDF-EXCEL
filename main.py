from pdfminer.high_level import extract_text
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

# Caminho para o arquivo PDF
pdf_path = "Relatório de Diagnóstico Mind The Bizz 2022.1.pdf"

# Extrai o texto de todas as páginas do arquivo PDF
text = extract_text(pdf_path)

# Divide o texto em páginas
pages = text.split('\x0c')  # Separa todas as páginas

# Cria um novo arquivo Excel
wb = Workbook()
ws = wb.active

# Adicionar labels das colunas
col_labels = [
    "Nome do projeto/negócio mentorado",
    "Setor/Segmento de atuação",
    "Responsável 1 | Profissão | E-mail | Telefone",
    "Responsável 2 | Profissão | E-mail | Telefone",
    "Local de Origem",
    "Nível de maturidade do projeto",
    "Resumo Diagnóstico",
    "Status de Desenvolvimento de Entregas",
    "Nome do mentor responsável pelo projeto/negócio"
]
# Configurar estilo das células dos labels das colunas
for col_num, label in enumerate(col_labels, start=1):
    cell = ws.cell(row=1, column=col_num, value=label.strip())  # Remover espaços em branco
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cor cinza
    ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = len(label.strip()) + 5
    
# Percorrer as páginas e preencher as células correspondentes
row_num = 2  # Inicia na segunda linha (após os labels das colunas)
for page_num, page in enumerate(pages, start=1):
    
    # Verificar se é a primeira página para ignorar completamente
    if page_num == 1:
        continue
    
    # Verificar se é uma página subsequente para remover as primeiras frases
    if page_num >= 2:
        
        lines = page.split('\n')[8:]  # Remove as primeiras linhas
    else:
        lines = page.split('\n')
    phrases = [
        "Nome do Projeto",
        "Segmento de atuação",
        "Responsável 1 | Profissão | E-mail | Telefone",
        "Responsável 2 | Profissão | E-mail | Telefone",
        "Local de origem",
        "Nível de maturidade do projeto",
        "Ideação",
        "Ideação + Protótipo",
        "Relatório Intermediário de Acompanhamento | Resumo",
        "Relatório de Diagnóstico | Resumo",
        "Status de Desenvolvimento de Entregas | Mind The Bizz",
        "Mentor(a):"
    ]
    
    # Remover as palavras indesejadas que estão nas frases
    lines = [line for line in lines if not any(word in line for word in phrases)]
    
    # Preencher as células da linha com as informações da página
    for col_num, line in enumerate(lines, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=line.strip())  # Remover espaços em branco
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row_num += 1
    
# Salvar o arquivo Excel
wb.save("teste.xlsx")